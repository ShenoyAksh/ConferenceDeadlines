#!/usr/bin/env python3
"""Scrape deadline pages for a CORE area sheet.

The script reads one area worksheet from CORE_by_area.xlsx, searches the web
for each conference, and writes a new workbook with these columns appended:

    URL, Abstract Deadline, Submission Deadline, Next Deadline, Countdown,
    Page Limit / Format

By default it looks for 2027 deadlines first, then falls back to 2026. The URL
written to the workbook is the page where the best deadline evidence was found,
or the best matching conference page when no deadline text could be extracted.

CORE C rows are skipped by default; the default included ranks are A*, A, and B.
The output sheet is sorted by upcoming deadline unless --no-sort is used.
"""

from __future__ import annotations

import argparse
import html
import json
import re
import sys
import time
from dataclasses import dataclass
from datetime import date, datetime, time as datetime_time
from html.parser import HTMLParser
from pathlib import Path
from typing import Iterable
from urllib import error, parse, request

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


USER_AGENT = (
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
    "AppleWebKit/537.36 (KHTML, like Gecko) "
    "Chrome/125.0 Safari/537.36"
)

MONTHS = {
    "jan": 1,
    "january": 1,
    "feb": 2,
    "february": 2,
    "mar": 3,
    "march": 3,
    "apr": 4,
    "april": 4,
    "may": 5,
    "jun": 6,
    "june": 6,
    "jul": 7,
    "july": 7,
    "aug": 8,
    "august": 8,
    "sep": 9,
    "sept": 9,
    "september": 9,
    "oct": 10,
    "october": 10,
    "nov": 11,
    "november": 11,
    "dec": 12,
    "december": 12,
}

MONTH_PATTERN = (
    r"Jan(?:uary)?|Feb(?:ruary)?|Mar(?:ch)?|Apr(?:il)?|May|Jun(?:e)?|"
    r"Jul(?:y)?|Aug(?:ust)?|Sep(?:t(?:ember)?)?|Oct(?:ober)?|"
    r"Nov(?:ember)?|Dec(?:ember)?"
)

DATE_RE = re.compile(
    rf"""
    (?P<date>
        (?:{MONTH_PATTERN})\.?\s+\d{{1,2}}(?:st|nd|rd|th)?
            (?:\s*[-–]\s*\d{{1,2}}(?:st|nd|rd|th)?)?,?\s+\d{{4}}
        |
        \d{{1,2}}(?:st|nd|rd|th)?\s+(?:{MONTH_PATTERN})\.?,?\s+\d{{4}}
        |
        \d{{4}}[-/]\d{{1,2}}[-/]\d{{1,2}}
        |
        \d{{1,2}}[-/]\d{{1,2}}[-/]\d{{4}}
    )
    """,
    re.IGNORECASE | re.VERBOSE,
)

BLOCK_TAGS = {
    "address",
    "article",
    "aside",
    "br",
    "dd",
    "div",
    "dl",
    "dt",
    "figcaption",
    "footer",
    "h1",
    "h2",
    "h3",
    "h4",
    "h5",
    "h6",
    "header",
    "li",
    "main",
    "nav",
    "ol",
    "p",
    "pre",
    "section",
    "table",
    "td",
    "th",
    "tr",
    "ul",
}

SKIP_TAGS = {"script", "style", "noscript", "svg", "canvas"}

SEARCH_HOSTS = {
    "duckduckgo.com",
    "www.duckduckgo.com",
    "lite.duckduckgo.com",
    "html.duckduckgo.com",
}

SKIP_HOST_FRAGMENTS = (
    "facebook.com",
    "instagram.com",
    "linkedin.com",
    "youtube.com",
    "youtu.be",
    "twitter.com",
    "x.com",
    "reddit.com",
    "semanticscholar.org",
    "dblp.org",
    "portal.core.edu.au",
    "core.edu.au",
    "getpaperpilot.com",
    "wikipedia.org",
)

LOW_PRIORITY_HOST_FRAGMENTS = (
    "wikicfp.com",
    "conference-service.com",
    "easychair.org",
    "openreview.net",
    "edas.info",
)

SKIP_PATH_SUFFIXES = (
    ".7z",
    ".avi",
    ".bmp",
    ".csv",
    ".doc",
    ".docx",
    ".gif",
    ".gz",
    ".jpg",
    ".jpeg",
    ".mov",
    ".mp3",
    ".mp4",
    ".pdf",
    ".png",
    ".ppt",
    ".pptx",
    ".tar",
    ".tgz",
    ".txt",
    ".xls",
    ".xlsx",
    ".zip",
)

RELEVANT_LINK_TERMS = (
    "author instruction",
    "author guideline",
    "important date",
    "deadline",
    "format",
    "instruction",
    "submission",
    "call for paper",
    "call-for-paper",
    "cfp",
    "author",
    "dates",
    "program",
    "template",
)

NEGATIVE_DEADLINE_TERMS = (
    "acceptance notification",
    "author notification",
    "camera-ready",
    "camera ready",
    "conference date",
    "conference dates",
    "early registration",
    "final version",
    "notification",
    "poster",
    "proceedings",
    "registration",
    "rebuttal",
    "tutorial proposal",
    "workshop proposal",
)

DEADLINE_HINT_TERMS = (
    "abstract",
    "article",
    "closes",
    "deadline",
    "due",
    "manuscript",
    "paper",
    "submit",
    "submission",
    "submissions",
)

PAGE_INFO_HINT_TERMS = (
    "aaai",
    "acm",
    "author",
    "ceur",
    "column",
    "format",
    "ieee",
    "instructions",
    "latex",
    "length",
    "lipics",
    "lncs",
    "page",
    "pages",
    "paper",
    "springer",
    "style",
    "submission",
    "template",
    "word",
)

PAGE_RANGE_RE = re.compile(
    r"\b(?P<low>\d{1,2})\s*(?:-|–|to|and)\s*(?P<high>\d{1,2})\s+(?:pages?|pp\.)\b",
    re.IGNORECASE,
)

PAGE_LIMIT_RE = re.compile(
    r"""
    (?:
        (?:
            at\s+most|maximum(?:\s+of)?|up\s+to|no\s+more\s+than|
            not\s+exceed(?:ing)?|limited\s+to|limit(?:ed)?\s+(?:of\s+)?|
            length\s+(?:is|of)?
        )
        \s+
    )?
    (?P<pages>\d{1,2})
    \s*
    (?:pages?|pp\.)
    (?:
        \s+(?:including|excluding)\s+(?:references?|appendices|bibliography|supplementary\s+material)
        |
        \s+(?:plus|for|maximum|max|limit|long|in\s+total|of\s+content)
    )?
    """,
    re.IGNORECASE | re.VERBOSE,
)

FORMAT_PATTERNS = (
    (re.compile(r"\b(?:Springer\s+)?LNCS(?:\s+(?:format|style|template))?\b", re.I), "LNCS format"),
    (re.compile(r"\bLecture Notes in Computer Science\b", re.I), "LNCS format"),
    (re.compile(r"\bLIPIcs(?:\s+(?:format|style|template))?\b", re.I), "LIPIcs format"),
    (re.compile(r"\bACM\s+(?:SIGCONF|Primary Article Template|format|template|style)\b", re.I), "ACM format"),
    (re.compile(r"\bIEEE(?:\s+conference)?\s+(?:format|template|style|two-column|2-column)\b", re.I), "IEEE format"),
    (re.compile(r"\bCEUR(?:-ART|\s+Workshop Proceedings)?(?:\s+(?:format|template|style))?\b", re.I), "CEUR format"),
    (re.compile(r"\bUSENIX\s+(?:format|template|style)\b", re.I), "USENIX format"),
    (re.compile(r"\bAAAI\s+(?:format|template|style)\b", re.I), "AAAI format"),
    (re.compile(r"\bIJCAI\s+(?:format|template|style)\b", re.I), "IJCAI format"),
    (re.compile(r"\b(?:single|double|two|2)[-\s]column\s+(?:format|style|template)?\b", re.I), None),
)


@dataclass(frozen=True)
class Link:
    url: str
    text: str


@dataclass(frozen=True)
class SearchHit:
    url: str
    title: str
    score: int


@dataclass(frozen=True)
class DeadlineCandidate:
    kind: str
    date_text: str
    year: str
    url: str
    context: str
    score: int
    parsed_date: date | None


@dataclass(frozen=True)
class PageInfoCandidate:
    kind: str
    value: str
    url: str
    context: str
    score: int


@dataclass(frozen=True)
class EnrichmentResult:
    url: str = ""
    abstract_deadline: str = ""
    submission_deadline: str = ""
    page_info: str = ""
    year: str = ""
    score: int = 0


class TextAndLinkParser(HTMLParser):
    def __init__(self, base_url: str):
        super().__init__(convert_charrefs=True)
        self.base_url = base_url
        self.text_parts: list[str] = []
        self.links: list[Link] = []
        self.skip_depth = 0
        self.current_link_href: str | None = None
        self.current_link_text: list[str] = []

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        tag = tag.lower()
        if tag in SKIP_TAGS:
            self.skip_depth += 1
            return

        if self.skip_depth:
            return

        if tag in BLOCK_TAGS:
            self.text_parts.append("\n")

        if tag == "a":
            href = dict(attrs).get("href")
            if href:
                self.current_link_href = parse.urljoin(self.base_url, href)
                self.current_link_text = []

    def handle_endtag(self, tag: str) -> None:
        tag = tag.lower()
        if tag in SKIP_TAGS and self.skip_depth:
            self.skip_depth -= 1
            return

        if self.skip_depth:
            return

        if tag == "a" and self.current_link_href:
            text = clean_space(" ".join(self.current_link_text))
            self.links.append(Link(self.current_link_href, text))
            self.current_link_href = None
            self.current_link_text = []

        if tag in BLOCK_TAGS:
            self.text_parts.append("\n")

    def handle_data(self, data: str) -> None:
        if self.skip_depth:
            return

        if data:
            self.text_parts.append(data)
            if self.current_link_href:
                self.current_link_text.append(data)

    @property
    def text(self) -> str:
        lines = [clean_space(part) for part in "".join(self.text_parts).splitlines()]
        return "\n".join(line for line in lines if line)


class HttpClient:
    def __init__(
        self,
        cache_path: Path | None,
        timeout: int,
        delay: float,
        max_bytes: int,
        verbose: bool,
    ) -> None:
        self.cache_path = cache_path
        self.timeout = timeout
        self.delay = delay
        self.max_bytes = max_bytes
        self.verbose = verbose
        self.last_fetch_time = 0.0
        self.cache: dict[str, dict[str, str | int]] = {}

        if cache_path and cache_path.exists():
            try:
                self.cache = json.loads(cache_path.read_text(encoding="utf-8"))
            except (OSError, json.JSONDecodeError):
                self.cache = {}

    def save_cache(self) -> None:
        if not self.cache_path:
            return

        self.cache_path.write_text(
            json.dumps(self.cache, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )

    def fetch(self, url: str) -> str:
        cached = self.cache.get(url)
        if cached and cached.get("ok") == 1:
            return str(cached.get("body", ""))

        now = time.monotonic()
        wait = self.delay - (now - self.last_fetch_time)
        if wait > 0:
            time.sleep(wait)

        if self.verbose:
            print(f"    fetch {url}")

        req = request.Request(
            url,
            headers={
                "User-Agent": USER_AGENT,
                "Accept": "text/html,application/xhtml+xml,text/plain;q=0.9,*/*;q=0.8",
                "Accept-Language": "en-US,en;q=0.9",
            },
        )

        try:
            with request.urlopen(req, timeout=self.timeout) as response:
                content_type = response.headers.get_content_type()
                charset = response.headers.get_content_charset() or "utf-8"
                raw = response.read(self.max_bytes)
        except (error.HTTPError, error.URLError, TimeoutError, OSError) as exc:
            self.cache[url] = {"ok": 0, "error": str(exc)}
            self.last_fetch_time = time.monotonic()
            return ""

        self.last_fetch_time = time.monotonic()

        if not (
            content_type.startswith("text/")
            or content_type in {"application/xhtml+xml", "application/xml"}
        ):
            self.cache[url] = {"ok": 0, "error": f"unsupported content-type {content_type}"}
            return ""

        body = raw.decode(charset, errors="replace")
        self.cache[url] = {"ok": 1, "body": body}
        return body


def clean_space(value: object) -> str:
    return re.sub(r"\s+", " ", html.unescape(str(value or ""))).strip()


def simplify_conference_name(name: str) -> str:
    name = re.sub(r";.*$", "", name)
    name = re.sub(r"\([^)]*(?:was|changed|duplicate|removed)[^)]*\)", "", name, flags=re.I)
    return clean_space(name)


def parse_html(html_text: str, base_url: str) -> tuple[str, list[Link]]:
    parser = TextAndLinkParser(base_url)
    try:
        parser.feed(html_text)
    except Exception:
        pass
    return parser.text, parser.links


def canonical_url(url: str) -> str:
    parsed = parse.urlsplit(url)
    query_pairs = [
        (key, value)
        for key, value in parse.parse_qsl(parsed.query, keep_blank_values=True)
        if not key.lower().startswith("utm_")
    ]
    return parse.urlunsplit(
        (
            parsed.scheme,
            parsed.netloc.lower(),
            parsed.path.rstrip("/") or "/",
            parse.urlencode(query_pairs),
            "",
        )
    )


def unwrap_search_url(url: str) -> str:
    url = html.unescape(url)
    parsed = parse.urlsplit(url)

    if parsed.netloc.lower() in SEARCH_HOSTS or not parsed.netloc:
        params = parse.parse_qs(parsed.query)
        for key in ("uddg", "u", "url"):
            if key in params and params[key]:
                return params[key][0]

    return url


def host_matches(host: str, fragments: Iterable[str]) -> bool:
    host = host.lower()
    return any(host == fragment or host.endswith(f".{fragment}") for fragment in fragments)


def should_skip_url(url: str) -> bool:
    parsed = parse.urlsplit(url)
    if parsed.scheme not in {"http", "https"} or not parsed.netloc:
        return True

    host = parsed.netloc.lower()
    if host in SEARCH_HOSTS or host_matches(host, SKIP_HOST_FRAGMENTS):
        return True

    path = parsed.path.lower()
    return any(path.endswith(suffix) for suffix in SKIP_PATH_SUFFIXES)


def meaningful_name_tokens(name: str) -> list[str]:
    stop_words = {
        "acm",
        "and",
        "conference",
        "for",
        "ieee",
        "international",
        "of",
        "on",
        "symposium",
        "the",
        "workshop",
    }
    return [
        token
        for token in re.findall(r"[a-z0-9]{4,}", simplify_conference_name(name).casefold())
        if token not in stop_words
    ]


def token_in_text(token: str, text: str) -> bool:
    return bool(re.search(rf"(?<![a-z0-9]){re.escape(token)}(?![a-z0-9])", text))


def hit_matches_conference(url: str, title: str, name: str, acronym: str) -> bool:
    haystack = f"{url} {title}".casefold()
    if acronym and re.search(rf"(?<![a-z0-9]){re.escape(acronym.casefold())}(?![a-z0-9])", haystack):
        return True

    token_hits = sum(1 for token in meaningful_name_tokens(name) if token_in_text(token, haystack))
    return token_hits >= 2


def page_matches_conference(text: str, url: str, name: str, acronym: str) -> bool:
    haystack = f"{url} {text[:12000]}".casefold()
    if acronym and re.search(rf"(?<![a-z0-9]){re.escape(acronym.casefold())}(?![a-z0-9])", haystack):
        return True

    token_hits = sum(1 for token in meaningful_name_tokens(name) if token_in_text(token, haystack))
    return token_hits >= 2


def search_score(url: str, title: str, name: str, acronym: str, year: str) -> int:
    parsed = parse.urlsplit(url)
    host = parsed.netloc.lower()
    haystack = f"{url} {title}".casefold()
    score = 0

    if year in haystack:
        score += 40
    if acronym and acronym.casefold() in haystack:
        score += 20

    name_tokens = meaningful_name_tokens(name)
    score += min(20, sum(4 for token in name_tokens if token_in_text(token, haystack)))

    if any(term in haystack for term in RELEVANT_LINK_TERMS):
        score += 20
    if host_matches(host, LOW_PRIORITY_HOST_FRAGMENTS):
        score -= 20

    return score


def build_queries(name: str, acronym: str, year: str, max_queries: int) -> list[str]:
    simple_name = simplify_conference_name(name)
    queries = []

    if acronym:
        queries.extend(
            [
                f'"{acronym} {year}" "submission deadline"',
                f'"{acronym}" "{year}" "important dates"',
                f'"{acronym}" "{year}" "call for papers"',
            ]
        )

    queries.extend(
        [
            f'"{simple_name}" "{year}" "submission deadline"',
            f'"{simple_name}" "{year}" "abstract submission"',
            f'"{simple_name}" "{year}" "important dates"',
        ]
    )

    deduped = []
    seen = set()
    for query in queries:
        if query not in seen:
            deduped.append(query)
            seen.add(query)
    return deduped[:max_queries]


def search_web(
    client: HttpClient,
    query: str,
    name: str,
    acronym: str,
    year: str,
    max_results: int,
) -> list[SearchHit]:
    search_url = "https://lite.duckduckgo.com/lite/?" + parse.urlencode({"q": query})
    html_text = client.fetch(search_url)
    if not html_text:
        return []

    _, links = parse_html(html_text, search_url)
    hits: list[SearchHit] = []
    seen_urls: set[str] = set()

    for link in links:
        unwrapped = canonical_url(unwrap_search_url(link.url))
        if should_skip_url(unwrapped) or unwrapped in seen_urls:
            continue
        if not hit_matches_conference(unwrapped, link.text, name, acronym):
            continue

        score = search_score(unwrapped, link.text, name, acronym, year)
        if score < 0:
            continue

        hits.append(SearchHit(unwrapped, link.text, score))
        seen_urls.add(unwrapped)

    hits.sort(key=lambda hit: hit.score, reverse=True)
    return hits[:max_results]


def relevant_child_links(page_url: str, links: list[Link], year: str, limit: int) -> list[str]:
    base_host = parse.urlsplit(page_url).netloc.lower()
    scored: list[tuple[int, str]] = []
    seen: set[str] = {canonical_url(page_url)}

    for link in links:
        url = canonical_url(link.url)
        parsed = parse.urlsplit(url)
        if should_skip_url(url) or parsed.netloc.lower() != base_host or url in seen:
            continue

        haystack = f"{url} {link.text}".casefold()
        score = 0
        if year in haystack:
            score += 20
        for term in RELEVANT_LINK_TERMS:
            if term in haystack:
                score += 12

        if score:
            scored.append((score, url))
            seen.add(url)

    scored.sort(reverse=True)
    return [url for _, url in scored[:limit]]


def candidate_year(date_text: str, years: list[str]) -> str:
    for year in years:
        if re.search(rf"\b{re.escape(year)}\b", date_text):
            return year
    return ""


def parse_date_text(date_text: str) -> date | None:
    text = clean_space(date_text).replace(".", "")
    text = re.sub(r"(\d)(st|nd|rd|th)", r"\1", text, flags=re.I)
    text = re.sub(r"\s*[-–]\s*\d{1,2}\b", "", text)

    iso_match = re.fullmatch(r"(\d{4})[-/](\d{1,2})[-/](\d{1,2})", text)
    if iso_match:
        year, month, day = map(int, iso_match.groups())
        try:
            return date(year, month, day)
        except ValueError:
            return None

    month_first = re.fullmatch(
        rf"({MONTH_PATTERN})\s+(\d{{1,2}}),?\s+(\d{{4}})",
        text,
        flags=re.I,
    )
    if month_first:
        month_name, day, year = month_first.groups()
        try:
            return date(int(year), MONTHS[month_name.casefold()[:3]], int(day))
        except (KeyError, ValueError):
            return None

    day_first = re.fullmatch(
        rf"(\d{{1,2}})\s+({MONTH_PATTERN}),?\s+(\d{{4}})",
        text,
        flags=re.I,
    )
    if day_first:
        day, month_name, year = day_first.groups()
        try:
            return date(int(year), MONTHS[month_name.casefold()[:3]], int(day))
        except (KeyError, ValueError):
            return None

    numeric = re.fullmatch(r"(\d{1,2})[-/](\d{1,2})[-/](\d{4})", text)
    if numeric:
        first, second, year = map(int, numeric.groups())
        month, day = (second, first) if first > 12 else (first, second)
        try:
            return date(year, month, day)
        except ValueError:
            return None

    return None


def deadline_score(context: str, kind: str) -> int:
    lower = context.casefold()
    score = 0
    has_deadline_word = any(term in lower for term in ("deadline", "due", "closes"))
    has_submission_word = any(term in lower for term in ("submission", "submissions", "submit"))
    has_paper_word = any(
        term in lower
        for term in (
            "paper",
            "papers",
            "full paper",
            "research track",
            "technical track",
            "manuscript",
            "article",
        )
    )

    if kind == "abstract":
        if "abstract" not in lower:
            return 0
        score += 50
        if "abstract submission" in lower or "abstract deadline" in lower:
            score += 35
        if "abstract registration" in lower:
            score += 25
        if has_deadline_word or has_submission_word:
            score += 20

    elif kind == "submission":
        if "abstract" in lower and not has_paper_word:
            return 0
        if has_submission_word and (has_paper_word or "deadline" in lower):
            score += 45
        if "paper submission" in lower or "submission deadline" in lower:
            score += 35
        if "full paper" in lower or "regular paper" in lower:
            score += 25
        if has_deadline_word:
            score += 20
    else:
        return 0

    if any(term in lower for term in NEGATIVE_DEADLINE_TERMS):
        score -= 40
    if "extended" in lower:
        score += 8

    return max(score, 0)


def has_deadline_hint(text: str) -> bool:
    lower = text.casefold()
    return any(term in lower for term in DEADLINE_HINT_TERMS)


def has_negative_deadline_hint(text: str) -> bool:
    lower = text.casefold()
    return any(term in lower for term in NEGATIVE_DEADLINE_TERMS)


def iter_date_contexts(text: str) -> Iterable[tuple[str, str]]:
    lines = [clean_space(line) for line in text.splitlines() if clean_space(line)]
    for index, line in enumerate(lines):
        if not DATE_RE.search(line):
            continue

        previous_line = lines[index - 1] if index > 0 else ""
        next_line = lines[index + 1] if index + 1 < len(lines) else ""
        line_has_hint = has_deadline_hint(line)
        previous_has_hint = has_deadline_hint(previous_line)
        next_has_hint = has_deadline_hint(next_line)
        line_has_negative_hint = has_negative_deadline_hint(line)

        for match in DATE_RE.finditer(line):
            date_text = clean_space(match.group("date"))
            if line_has_negative_hint and not line_has_hint:
                context = line
            elif line_has_hint:
                start = max(0, match.start() - 100)
                end = min(len(line), match.end() + 100)
                context = line[start:end]
            elif previous_has_hint:
                context = f"{previous_line} {line}"
            elif next_has_hint:
                context = f"{line} {next_line}"
            else:
                context = f"{previous_line} {line} {next_line}"

            yield date_text, clean_space(context)


def context_mentions_different_date(context: str, date_text: str) -> bool:
    dates = [clean_space(match.group("date")) for match in DATE_RE.finditer(context)]
    return any(date != date_text for date in dates)


def extract_deadline_candidates(
    text: str,
    url: str,
    years: list[str],
) -> list[DeadlineCandidate]:
    candidates: list[DeadlineCandidate] = []
    seen: set[tuple[str, str, str, str]] = set()

    for date_text, context in iter_date_contexts(text):
        year = candidate_year(date_text, years)
        if not year:
            continue

        parsed_date = parse_date_text(date_text)

        for kind in ("abstract", "submission"):
            score = deadline_score(context, kind)
            if context_mentions_different_date(context, date_text):
                score -= 25
            if score <= 0:
                continue

            key = (kind, date_text, year, url)
            if key in seen:
                continue
            seen.add(key)

            candidates.append(
                DeadlineCandidate(
                    kind=kind,
                    date_text=date_text,
                    year=year,
                    url=url,
                    context=context,
                    score=score,
                    parsed_date=parsed_date,
                )
            )

    return candidates


def has_page_info_hint(text: str) -> bool:
    lower = text.casefold()
    return any(term in lower for term in PAGE_INFO_HINT_TERMS)


def page_info_contexts(text: str) -> Iterable[str]:
    lines = [clean_space(line) for line in text.splitlines() if clean_space(line)]
    for index, line in enumerate(lines):
        if not has_page_info_hint(line):
            continue

        start = max(0, index - 1)
        end = min(len(lines), index + 2)
        yield clean_space(" ".join(lines[start:end]))


def normalize_page_phrase(match: re.Match[str]) -> str:
    phrase = clean_space(match.group(0))
    pages = match.groupdict().get("pages")
    if pages and phrase == pages:
        return f"{pages} pages"
    return phrase


def page_info_score(context: str, kind: str) -> int:
    lower = context.casefold()
    score = 20

    if kind == "page_limit":
        score += 35
    elif kind == "format":
        score += 25

    for term in ("author", "instructions", "submission", "paper", "papers", "format", "template"):
        if term in lower:
            score += 10

    for term in ("full paper", "regular paper", "research paper", "technical paper"):
        if term in lower:
            score += 15

    for term in ("accepted", "camera-ready", "camera ready", "proceedings", "presentation"):
        if term in lower:
            score -= 12

    return max(score, 1)


def extract_page_info_candidates(text: str, url: str) -> list[PageInfoCandidate]:
    candidates: list[PageInfoCandidate] = []
    seen: set[tuple[str, str, str]] = set()

    for context in page_info_contexts(text):
        for match in PAGE_RANGE_RE.finditer(context):
            low = int(match.group("low"))
            high = int(match.group("high"))
            if low > high:
                low, high = high, low
            value = f"{low}-{high} pages"
            key = ("page_limit", value.casefold(), url)
            if key not in seen:
                seen.add(key)
                candidates.append(
                    PageInfoCandidate(
                        kind="page_limit",
                        value=value,
                        url=url,
                        context=context,
                        score=page_info_score(context, "page_limit") + 15,
                    )
                )

        for match in PAGE_LIMIT_RE.finditer(context):
            pages = int(match.group("pages"))
            if pages < 2 or pages > 50:
                continue

            value = normalize_page_phrase(match)
            key = ("page_limit", value.casefold(), url)
            if key not in seen:
                seen.add(key)
                candidates.append(
                    PageInfoCandidate(
                        kind="page_limit",
                        value=value,
                        url=url,
                        context=context,
                        score=page_info_score(context, "page_limit"),
                    )
                )

        for pattern, normalized_value in FORMAT_PATTERNS:
            for match in pattern.finditer(context):
                value = normalized_value or clean_space(match.group(0))
                key = ("format", value.casefold(), url)
                if key not in seen:
                    seen.add(key)
                    candidates.append(
                        PageInfoCandidate(
                            kind="format",
                            value=value,
                            url=url,
                            context=context,
                            score=page_info_score(context, "format"),
                        )
                    )

    return candidates


def best_page_info(candidates: list[PageInfoCandidate]) -> str:
    selected: list[str] = []

    for kind in ("page_limit", "format"):
        matching = [candidate for candidate in candidates if candidate.kind == kind]
        matching.sort(key=lambda candidate: (candidate.score, len(candidate.value)), reverse=True)
        if matching:
            value = matching[0].value
            if value.casefold() not in {item.casefold() for item in selected}:
                selected.append(value)

    return "; ".join(selected)


def best_candidate(
    candidates: list[DeadlineCandidate],
    kind: str,
    year: str,
) -> DeadlineCandidate | None:
    matching = [candidate for candidate in candidates if candidate.kind == kind and candidate.year == year]
    if not matching:
        return None

    matching.sort(
        key=lambda candidate: (
            candidate.score,
            candidate.parsed_date or date.min,
        ),
        reverse=True,
    )
    return matching[0]


def inspect_hit(
    client: HttpClient,
    hit: SearchHit,
    name: str,
    acronym: str,
    year: str,
    years: list[str],
    child_link_limit: int,
) -> EnrichmentResult:
    pages = [hit.url]
    all_candidates: list[DeadlineCandidate] = []
    all_page_info_candidates: list[PageInfoCandidate] = []
    best_url = hit.url

    html_text = client.fetch(hit.url)
    if html_text:
        text, links = parse_html(html_text, hit.url)
        if not page_matches_conference(text, hit.url, name, acronym):
            return EnrichmentResult(score=-1)
        all_candidates.extend(extract_deadline_candidates(text, hit.url, years))
        all_page_info_candidates.extend(extract_page_info_candidates(text, hit.url))
        pages.extend(relevant_child_links(hit.url, links, year, child_link_limit))

    for page_url in pages[1:]:
        page_html = client.fetch(page_url)
        if not page_html:
            continue
        text, _ = parse_html(page_html, page_url)
        if not page_matches_conference(text, page_url, name, acronym):
            continue
        all_candidates.extend(extract_deadline_candidates(text, page_url, years))
        all_page_info_candidates.extend(extract_page_info_candidates(text, page_url))

    abstract = best_candidate(all_candidates, "abstract", year)
    submission = best_candidate(all_candidates, "submission", year)
    page_info = best_page_info(all_page_info_candidates)

    score = hit.score
    if abstract:
        score += 120 + abstract.score
        best_url = abstract.url
    if submission:
        score += 120 + submission.score
        if not abstract or submission.score >= abstract.score:
            best_url = submission.url
    if page_info:
        score += 60

    if year in best_url:
        score += 15

    return EnrichmentResult(
        url=best_url,
        abstract_deadline=abstract.date_text if abstract else "",
        submission_deadline=submission.date_text if submission else "",
        page_info=page_info,
        year=year if abstract or submission else "",
        score=score,
    )


def enrich_conference(
    client: HttpClient,
    name: str,
    acronym: str,
    years: list[str],
    queries_per_year: int,
    max_search_results: int,
    pages_per_year: int,
    child_link_limit: int,
    verbose: bool,
) -> EnrichmentResult:
    best_by_year: dict[str, EnrichmentResult] = {}

    for year in years:
        hits_by_url: dict[str, SearchHit] = {}
        for query in build_queries(name, acronym, year, queries_per_year):
            if verbose:
                print(f"    search {query}")
            for hit in search_web(client, query, name, acronym, year, max_search_results):
                current = hits_by_url.get(hit.url)
                if current is None or hit.score > current.score:
                    hits_by_url[hit.url] = hit

        hits = sorted(hits_by_url.values(), key=lambda hit: hit.score, reverse=True)
        year_results: list[EnrichmentResult] = []
        for hit in hits[:pages_per_year]:
            result = inspect_hit(client, hit, name, acronym, year, years, child_link_limit)
            if result.score < 0:
                continue
            year_results.append(result)
            if result.abstract_deadline and result.submission_deadline and result.page_info:
                break

        if year_results:
            year_results.sort(key=lambda result: result.score, reverse=True)
            best_by_year[year] = year_results[0]

        best = best_by_year.get(year)
        if best and (best.abstract_deadline or best.submission_deadline):
            return best

    for year in years:
        best = best_by_year.get(year)
        if best and best.url:
            return best

    return EnrichmentResult()


def header_map(sheet) -> dict[str, int]:
    return {
        clean_space(cell.value).casefold(): cell.column
        for cell in sheet[1]
        if clean_space(cell.value)
    }


def ensure_column(sheet, name: str) -> int:
    headers = header_map(sheet)
    existing = headers.get(name.casefold())
    if existing:
        return existing

    column = sheet.max_column + 1
    cell = sheet.cell(row=1, column=column, value=name)
    cell.font = Font(bold=True)
    cell.fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    return column


def ensure_column_before(sheet, name: str, before_name: str) -> int:
    headers = header_map(sheet)
    existing = headers.get(name.casefold())
    if existing:
        return existing

    before_column = headers.get(before_name.casefold())
    if before_column:
        sheet.insert_cols(before_column)
        column = before_column
    else:
        column = sheet.max_column + 1

    cell = sheet.cell(row=1, column=column, value=name)
    cell.font = Font(bold=True)
    cell.fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    return column


def set_column_width(sheet, column: int, width: int) -> None:
    sheet.column_dimensions[get_column_letter(column)].width = width


def write_result(sheet, row: int, columns: dict[str, int], result: EnrichmentResult, overwrite: bool) -> None:
    values = {
        "URL": result.url,
        "Abstract Deadline": result.abstract_deadline,
        "Submission Deadline": result.submission_deadline,
        "Page Limit / Format": result.page_info,
    }

    for header, value in values.items():
        cell = sheet.cell(row=row, column=columns[header])
        if overwrite or not clean_space(cell.value):
            cell.value = value
            if header == "URL" and value:
                cell.hyperlink = value
                cell.style = "Hyperlink"


def next_deadline_from_values(abstract_deadline: str, submission_deadline: str) -> date | None:
    parsed_dates = [
        parsed
        for parsed in (
            parse_date_text(abstract_deadline),
            parse_date_text(submission_deadline),
        )
        if parsed is not None
    ]

    if not parsed_dates:
        return None

    today = date.today()
    upcoming_dates = [parsed for parsed in parsed_dates if parsed >= today]
    return min(upcoming_dates or parsed_dates)


def update_next_deadline(sheet, row: int, columns: dict[str, int]) -> date | None:
    next_deadline = next_deadline_from_values(
        clean_space(sheet.cell(row=row, column=columns["Abstract Deadline"]).value),
        clean_space(sheet.cell(row=row, column=columns["Submission Deadline"]).value),
    )
    cell = sheet.cell(row=row, column=columns["Next Deadline"])
    cell.value = next_deadline.isoformat() if next_deadline else ""
    return next_deadline


def format_countdown(next_deadline: date | None) -> str:
    if next_deadline is None:
        return ""

    now = datetime.now()
    deadline_end = datetime.combine(next_deadline, datetime_time.max.replace(microsecond=0))
    total_seconds = int((deadline_end - now).total_seconds())
    overdue = total_seconds < 0
    total_seconds = abs(total_seconds)

    days, remainder = divmod(total_seconds, 24 * 60 * 60)
    hours, remainder = divmod(remainder, 60 * 60)
    minutes = remainder // 60

    if days:
        value = f"{days} days {hours} hours"
    elif hours:
        value = f"{hours} hours {minutes} minutes"
    else:
        value = f"{minutes} minutes"

    return f"overdue by {value}" if overdue else value


def update_countdown(sheet, row: int, columns: dict[str, int], next_deadline: date | None = None) -> None:
    if next_deadline is None:
        next_deadline = parse_date_text(clean_space(sheet.cell(row=row, column=columns["Next Deadline"]).value))

    sheet.cell(row=row, column=columns["Countdown"]).value = format_countdown(next_deadline)


def update_timing_columns(sheet, row: int, columns: dict[str, int]) -> date | None:
    next_deadline = update_next_deadline(sheet, row, columns)
    update_countdown(sheet, row, columns, next_deadline)
    return next_deadline


def sort_by_upcoming_deadline(sheet, columns: dict[str, int]) -> None:
    today = date.today()
    rows: list[tuple[tuple[int, int, int], list[object], dict[int, str]]] = []

    for row_index in range(2, sheet.max_row + 1):
        next_deadline = update_timing_columns(sheet, row_index, columns)
        values = [sheet.cell(row=row_index, column=column).value for column in range(1, sheet.max_column + 1)]
        hyperlinks = {
            column: sheet.cell(row=row_index, column=column).hyperlink.target
            for column in range(1, sheet.max_column + 1)
            if sheet.cell(row=row_index, column=column).hyperlink
        }

        if next_deadline is None:
            sort_key = (2, 0, row_index)
        elif next_deadline >= today:
            sort_key = (0, next_deadline.toordinal(), row_index)
        else:
            sort_key = (1, -next_deadline.toordinal(), row_index)

        rows.append((sort_key, values, hyperlinks))

    rows.sort(key=lambda item: item[0])

    for target_row, (_, values, hyperlinks) in enumerate(rows, start=2):
        for column, value in enumerate(values, start=1):
            cell = sheet.cell(row=target_row, column=column)
            cell.value = value
            cell.hyperlink = None
            if column in hyperlinks:
                cell.hyperlink = hyperlinks[column]
                cell.style = "Hyperlink"


def rank_is_included(rank: str, include_ranks: list[str]) -> bool:
    normalized_rank = clean_space(rank).casefold()
    normalized_includes = {clean_space(value).casefold() for value in include_ranks}

    if "all" in normalized_includes:
        return True
    if normalized_rank in normalized_includes:
        return True

    aliases = {
        "australasian b": "b",
        "australasian c": "c",
    }
    return aliases.get(normalized_rank) in normalized_includes


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Scrape one CORE area sheet for conference URLs and submission deadlines."
    )
    parser.add_argument(
        "area",
        nargs="?",
        default="4612",
        help="area code / worksheet to enrich, default: 4612",
    )
    parser.add_argument(
        "-i",
        "--input",
        default="CORE_by_area.xlsx",
        help="input workbook, default: CORE_by_area.xlsx",
    )
    parser.add_argument(
        "-o",
        "--output",
        default="deadline.xlsx",
        help="output workbook, default: deadline.xlsx",
    )
    parser.add_argument(
        "--sheet",
        default="",
        help="worksheet to enrich; overrides the positional area code",
    )
    parser.add_argument(
        "--include-ranks",
        nargs="+",
        default=["A*", "A", "B"],
        help="CORE ranks to scrape, default: A* A B. Use ALL to scrape every row.",
    )
    parser.add_argument(
        "--years",
        nargs="+",
        default=["2027", "2026"],
        help="conference years to try in order, default: 2027 2026",
    )
    parser.add_argument("--limit", type=int, default=0, help="limit rows processed, default: all")
    parser.add_argument("--start-row", type=int, default=2, help="first worksheet row to process")
    parser.add_argument("--overwrite", action="store_true", help="replace existing URL/deadline values")
    parser.add_argument("--dry-run", action="store_true", help="do not write the output workbook")
    parser.add_argument(
        "--no-sort",
        action="store_true",
        help="keep original sheet order instead of sorting by upcoming deadline",
    )
    parser.add_argument("--timeout", type=int, default=15, help="HTTP timeout in seconds")
    parser.add_argument("--delay", type=float, default=0.75, help="delay between HTTP requests")
    parser.add_argument("--max-bytes", type=int, default=2_000_000, help="maximum bytes read per page")
    parser.add_argument(
        "--queries-per-year",
        type=int,
        default=4,
        help="search queries per conference/year",
    )
    parser.add_argument(
        "--max-search-results",
        type=int,
        default=6,
        help="search results kept from each query",
    )
    parser.add_argument(
        "--pages-per-year",
        type=int,
        default=6,
        help="top search-result pages inspected per year",
    )
    parser.add_argument(
        "--child-links",
        type=int,
        default=3,
        help="same-site deadline/CFP links followed from each result page",
    )
    parser.add_argument(
        "--cache",
        default=".core_deadline_scrape_cache.json",
        help="HTTP cache JSON path; use empty string to disable",
    )
    parser.add_argument("--verbose", action="store_true", help="print search and fetch details")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    input_path = Path(args.input)
    output_path = Path(args.output)
    cache_path = Path(args.cache) if args.cache else None
    years = [str(year) for year in args.years]
    target_sheet = args.sheet or args.area

    workbook = load_workbook(input_path)
    if target_sheet not in workbook.sheetnames:
        print(f"Workbook has no sheet named {target_sheet!r}", file=sys.stderr)
        return 1

    sheet = workbook[target_sheet]
    headers = header_map(sheet)
    try:
        conference_col = headers["conference"]
        acronym_col = headers["acronym"]
        rank_col = headers["rank"]
    except KeyError as exc:
        print(f"Missing required column: {exc.args[0]}", file=sys.stderr)
        return 1

    for column_name in ("URL", "Abstract Deadline", "Submission Deadline", "Next Deadline"):
        ensure_column(sheet, column_name)
    ensure_column(sheet, "Page Limit / Format")
    ensure_column_before(sheet, "Countdown", "Page Limit / Format")

    headers = header_map(sheet)
    output_columns = {
        column_name: headers[column_name.casefold()]
        for column_name in (
            "URL",
            "Abstract Deadline",
            "Submission Deadline",
            "Next Deadline",
            "Countdown",
            "Page Limit / Format",
        )
    }
    set_column_width(sheet, output_columns["URL"], 50)
    set_column_width(sheet, output_columns["Abstract Deadline"], 22)
    set_column_width(sheet, output_columns["Submission Deadline"], 24)
    set_column_width(sheet, output_columns["Next Deadline"], 18)
    set_column_width(sheet, output_columns["Countdown"], 22)
    set_column_width(sheet, output_columns["Page Limit / Format"], 32)

    client = HttpClient(
        cache_path=cache_path,
        timeout=args.timeout,
        delay=args.delay,
        max_bytes=args.max_bytes,
        verbose=args.verbose,
    )

    processed = 0
    skipped_rank = 0
    max_row = sheet.max_row

    try:
        for row in range(args.start_row, max_row + 1):
            if args.limit and processed >= args.limit:
                break

            existing_values = [
                clean_space(sheet.cell(row=row, column=column).value)
                for column in (
                    output_columns["URL"],
                    output_columns["Abstract Deadline"],
                    output_columns["Submission Deadline"],
                    output_columns["Page Limit / Format"],
                )
            ]
            if not args.overwrite and all(existing_values):
                update_timing_columns(sheet, row, output_columns)
                continue

            name = clean_space(sheet.cell(row=row, column=conference_col).value)
            acronym = clean_space(sheet.cell(row=row, column=acronym_col).value)
            if not name:
                continue

            rank = clean_space(sheet.cell(row=row, column=rank_col).value)
            if not rank_is_included(rank, args.include_ranks):
                skipped_rank += 1
                if args.verbose:
                    print(f"skip row {row}: rank={rank or '-'}")
                continue

            processed += 1
            label = f"{acronym} - {name}" if acronym else name
            print(f"[{processed}] row {row}: {label}")

            result = enrich_conference(
                client=client,
                name=name,
                acronym=acronym,
                years=years,
                queries_per_year=args.queries_per_year,
                max_search_results=args.max_search_results,
                pages_per_year=args.pages_per_year,
                child_link_limit=args.child_links,
                verbose=args.verbose,
            )

            if result.url:
                found_bits = [
                    f"year={result.year or 'page-only'}",
                    f"abstract={result.abstract_deadline or '-'}",
                    f"submission={result.submission_deadline or '-'}",
                    f"pages/format={result.page_info or '-'}",
                ]
                print("    " + ", ".join(found_bits))
                print(f"    {result.url}")
            else:
                print("    no matching page found")

            write_result(sheet, row, output_columns, result, overwrite=args.overwrite)
            update_timing_columns(sheet, row, output_columns)

            if cache_path:
                client.save_cache()

    finally:
        if cache_path:
            client.save_cache()

    if args.dry_run:
        print(f"Dry run complete; no workbook written. Skipped {skipped_rank} row(s) by rank.")
        return 0

    for worksheet in list(workbook.worksheets):
        if worksheet.title != target_sheet:
            workbook.remove(worksheet)

    if not args.no_sort:
        sort_by_upcoming_deadline(sheet, output_columns)

    sheet.auto_filter.ref = sheet.dimensions
    workbook.save(output_path)
    print(
        f"Wrote enriched workbook to {output_path} "
        f"with sheet {target_sheet!r}; skipped {skipped_rank} row(s) by rank."
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
