#!/usr/bin/env python3
"""Create an Excel workbook with one sheet per CORE area code.

The bundled CORE.csv has no header row and uses this layout:

    id, conference name, acronym, source, rank, listed, area1, area2, area3

Each conference is written to every area sheet it belongs to, ordered by CORE
rank with the strongest rankings first.
"""

from __future__ import annotations

import argparse
import csv
import re
import sys
from collections import defaultdict
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:  # pragma: no cover - depends on local environment
    print(
        "This script needs openpyxl. Install it with: python3 -m pip install openpyxl",
        file=sys.stderr,
    )
    raise SystemExit(1)


HEADERS = [
    "CORE ID",
    "Conference",
    "Acronym",
    "Source",
    "Rank",
    "Listed",
    "Area 1",
    "Area 2",
    "Area 3",
]

RANK_ORDER = {
    "a*": 0,
    "a": 1,
    "b": 2,
    "australasian b": 2,
    "c": 3,
    "australasian c": 3,
}

LOWER_RANK_PREFIXES = (
    ("national", 4),
    ("regional", 4),
    ("multiconference", 5),
    ("journal published", 6),
    ("unranked", 7),
)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Split a CORE CSV into an Excel workbook with one sheet per area."
    )
    parser.add_argument(
        "input",
        nargs="?",
        default="CORE.csv",
        help="input CORE CSV file, default: CORE.csv",
    )
    parser.add_argument(
        "-o",
        "--output",
        default="CORE_by_area.xlsx",
        help="output Excel workbook, default: CORE_by_area.xlsx",
    )
    parser.add_argument(
        "--area-start-column",
        type=int,
        default=7,
        help="1-based column where area codes start, default: 7",
    )
    return parser.parse_args()


def rank_sort_key(row: list[str]) -> tuple[int, str, str]:
    rank = row[4].strip().casefold()
    if rank in RANK_ORDER:
        rank_score = RANK_ORDER[rank]
    else:
        rank_score = 8
        for prefix, score in LOWER_RANK_PREFIXES:
            if rank.startswith(prefix):
                rank_score = score
                break

    conference = row[1].strip().casefold()
    acronym = row[2].strip().casefold()
    return rank_score, conference, acronym


def safe_sheet_title(raw_title: str, used_titles: set[str]) -> str:
    title = re.sub(r"[\[\]:*?/\\]", "_", raw_title.strip())[:31] or "Area"
    candidate = title
    suffix = 2

    while candidate in used_titles:
        suffix_text = f"_{suffix}"
        candidate = f"{title[:31 - len(suffix_text)]}{suffix_text}"
        suffix += 1

    used_titles.add(candidate)
    return candidate


def read_rows(input_path: Path) -> list[list[str]]:
    with input_path.open(newline="", encoding="utf-8-sig") as csv_file:
        rows = list(csv.reader(csv_file))

    if not rows:
        raise ValueError(f"{input_path} is empty")

    expected_width = len(HEADERS)
    bad_rows = [index for index, row in enumerate(rows, start=1) if len(row) != expected_width]
    if bad_rows:
        sample = ", ".join(str(row_number) for row_number in bad_rows[:10])
        raise ValueError(
            f"Expected {expected_width} columns, but found different widths on row(s): {sample}"
        )

    return rows


def collect_by_area(rows: list[list[str]], area_start_index: int) -> dict[str, list[list[str]]]:
    by_area: dict[str, list[list[str]]] = defaultdict(list)

    for row in rows:
        seen_for_row: set[str] = set()
        for area in row[area_start_index:]:
            area = area.strip()
            if area and area not in seen_for_row:
                by_area[area].append(row)
                seen_for_row.add(area)

    return dict(sorted(by_area.items()))


def set_column_widths(sheet) -> None:
    for column in sheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column)
        width = min(max(max_length + 2, 10), 70)
        sheet.column_dimensions[get_column_letter(column[0].column)].width = width


def write_sheet(workbook: Workbook, title: str, rows: list[list[str]]) -> None:
    sheet = workbook.create_sheet(title=title)
    sheet.append(HEADERS)

    for row in sorted(rows, key=rank_sort_key):
        sheet.append([cell.strip() for cell in row])

    header_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    set_column_widths(sheet)


def main() -> int:
    args = parse_args()
    input_path = Path(args.input)
    output_path = Path(args.output)
    area_start_index = args.area_start_column - 1

    if area_start_index < 0:
        raise ValueError("--area-start-column must be 1 or greater")

    rows = read_rows(input_path)
    by_area = collect_by_area(rows, area_start_index)

    workbook = Workbook()
    workbook.remove(workbook.active)
    used_titles: set[str] = set()

    for area, area_rows in by_area.items():
        sheet_title = safe_sheet_title(area, used_titles)
        write_sheet(workbook, sheet_title, area_rows)

    workbook.save(output_path)
    print(f"Wrote {len(by_area)} area sheet(s) to {output_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
