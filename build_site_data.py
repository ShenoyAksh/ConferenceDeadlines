#!/usr/bin/env python3
"""Convert the final deadline workbook into data for the static website."""

from __future__ import annotations

import argparse
import json
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


def clean(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Build site/data.js from a deadline workbook.")
    parser.add_argument(
        "input",
        nargs="?",
        default="deadline.xlsx",
        help="input deadline workbook, default: deadline.xlsx",
    )
    parser.add_argument(
        "-o",
        "--output",
        default="site/data.js",
        help="output JavaScript data file, default: site/data.js",
    )
    return parser.parse_args()


def workbook_to_data(input_path: Path) -> dict[str, object]:
    workbook = load_workbook(input_path, read_only=True, data_only=True)
    sheets = []

    for worksheet in workbook.worksheets:
        headers = [clean(worksheet.cell(row=1, column=column).value) for column in range(1, worksheet.max_column + 1)]
        rows = []

        for row_index in range(2, worksheet.max_row + 1):
            row = {
                headers[column_index - 1]: clean(worksheet.cell(row=row_index, column=column_index).value)
                for column_index in range(1, worksheet.max_column + 1)
                if headers[column_index - 1]
            }
            if any(row.values()):
                rows.append(row)

        sheets.append(
            {
                "name": worksheet.title,
                "headers": headers,
                "rows": rows,
            }
        )

    return {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "source_workbook": str(input_path),
        "sheets": sheets,
    }


def main() -> int:
    args = parse_args()
    input_path = Path(args.input)
    output_path = Path(args.output)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    data = workbook_to_data(input_path)
    payload = json.dumps(data, ensure_ascii=False, indent=2)
    output_path.write_text(f"window.CORE_DEADLINE_DATA = {payload};\n", encoding="utf-8")

    row_count = sum(len(sheet["rows"]) for sheet in data["sheets"])
    print(f"Wrote {row_count} row(s) from {input_path} to {output_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
